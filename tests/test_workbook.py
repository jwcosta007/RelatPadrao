"""
Testes de integração sobre o workbook gerado pelo ETL.
Lê o arquivo xlsx mais recente em relatorios/ — skip se não existir.
Para testes com dados reais do piloto, usar o marker @dados_disponiveis.
"""
from pathlib import Path
import pytest
import openpyxl

from conftest import dados_disponiveis

BASE_DIR   = Path(__file__).parent.parent
RELAT_DIR  = BASE_DIR / "relatorios"

ABAS_ESPERADAS = [
    "DRE Gerencial", "DFC", "KPIs", "f_Base",
    "Lista", "f_Erros", "f_SaldoBancos", "cad_cliente", "check",
]


def _ultimo_relatorio():
    """Retorna o xlsx mais recente em relatorios/, ou None."""
    xlsxs = sorted(RELAT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsxs[0] if xlsxs else None


def _skip_sem_relatorio():
    return pytest.mark.skipif(
        _ultimo_relatorio() is None,
        reason="Nenhum relatório gerado em relatorios/ — execute etl.py AB primeiro"
    )


@pytest.fixture(scope="module")
def workbook():
    """Carrega o xlsx mais recente para todos os testes do módulo."""
    path = _ultimo_relatorio()
    if path is None:
        pytest.skip("Nenhum relatório encontrado em relatorios/")
    return openpyxl.load_workbook(path, data_only=True)


# ─────────────────────────────────────────────────────────────────────────────
# Estrutura do workbook
# ─────────────────────────────────────────────────────────────────────────────

class TestEstrutura:

    def test_todas_as_abas_presentes(self, workbook):
        abas = workbook.sheetnames
        for aba in ABAS_ESPERADAS:
            assert aba in abas, f"Aba '{aba}' não encontrada no workbook"

    def test_numero_de_abas(self, workbook):
        assert len(workbook.sheetnames) == len(ABAS_ESPERADAS)

    def test_nomes_de_abas_sem_extras(self, workbook):
        extras = set(workbook.sheetnames) - set(ABAS_ESPERADAS)
        assert extras == set(), f"Abas inesperadas: {extras}"


# ─────────────────────────────────────────────────────────────────────────────
# f_Base
# ─────────────────────────────────────────────────────────────────────────────

class TestFBase:

    def test_f_base_tem_34_colunas(self, workbook):
        ws = workbook["f_Base"]
        # Lê cabeçalho da tabela (linha 1)
        cabecalho = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        cabecalho = [c for c in cabecalho if c is not None]
        assert len(cabecalho) == 34, f"f_Base tem {len(cabecalho)} colunas, esperado 34"

    def test_f_base_tem_linhas_de_dados(self, workbook):
        ws = workbook["f_Base"]
        # Linha 1 = cabeçalho, linha 2+ = dados
        assert ws.max_row > 1, "f_Base não tem linhas de dados"

    def test_f_base_primeira_coluna_e_tipo_registro(self, workbook):
        ws = workbook["f_Base"]
        assert ws.cell(row=1, column=1).value == "tipo_registro"


# ─────────────────────────────────────────────────────────────────────────────
# f_Erros
# ─────────────────────────────────────────────────────────────────────────────

class TestFErros:

    def test_f_erros_sem_erros_tecnicos(self, workbook):
        ws = workbook["f_Erros"]
        # Linha 1 = cabeçalho; dados começam na linha 2
        dados = [
            ws.cell(row=r, column=7).value  # coluna 'motivo'
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None
        ]
        erros_tecnicos = [m for m in dados if m and "Sem mapeamento" not in str(m)]
        assert erros_tecnicos == [], f"Erros técnicos encontrados: {erros_tecnicos}"

    def test_f_erros_sem_sem_mapa(self, workbook):
        ws = workbook["f_Erros"]
        dados = [
            ws.cell(row=r, column=7).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None
        ]
        sem_mapa = [m for m in dados if m and "Sem mapeamento" in str(m)]
        assert sem_mapa == [], f"{len(sem_mapa)} lançamento(s) sem mapeamento no MapaAloc"


# ─────────────────────────────────────────────────────────────────────────────
# DRE Gerencial
# ─────────────────────────────────────────────────────────────────────────────

class TestDRE:

    def test_dre_tem_linhas_de_dados(self, workbook):
        ws = workbook["DRE Gerencial"]
        # Dados começam na linha DATA_ROW = 8
        valores = [ws.cell(row=r, column=2).value for r in range(8, ws.max_row + 1)]
        valores = [v for v in valores if v is not None]
        assert len(valores) > 0, "DRE Gerencial não tem linhas de dados"

    def test_dre_tem_kpi_receita_liquida(self, workbook):
        ws = workbook["DRE Gerencial"]
        labels = [ws.cell(row=r, column=2).value for r in range(8, ws.max_row + 1)]
        assert "RECEITA LÍQUIDA" in labels, "KPI 'RECEITA LÍQUIDA' não encontrado no DRE"


# ─────────────────────────────────────────────────────────────────────────────
# DFC
# ─────────────────────────────────────────────────────────────────────────────

class TestDFC:

    def test_dfc_tem_caixa_inicio(self, workbook):
        ws = workbook["DFC"]
        labels = [ws.cell(row=r, column=2).value for r in range(8, ws.max_row + 1)]
        assert "CAIXA - INÍCIO DO MÊS" in labels

    def test_dfc_tem_caixa_fim(self, workbook):
        ws = workbook["DFC"]
        labels = [ws.cell(row=r, column=2).value for r in range(8, ws.max_row + 1)]
        assert "CAIXA - FIM DO MÊS" in labels

    def test_dfc_tem_fluxo_de_caixa(self, workbook):
        ws = workbook["DFC"]
        labels = [ws.cell(row=r, column=2).value for r in range(8, ws.max_row + 1)]
        assert "FLUXO DE CAIXA" in labels
