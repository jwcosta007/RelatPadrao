"""
builder.py — Constrói estrutura do workbook AZ Resultados a partir do MapaAloc.
Chamado a cada carga pelo etl_ab.py (builder roda sempre, sem diff).
"""

import re
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import (
    AbsoluteAnchor, XDRPoint2D, XDRPositiveSize2D, pixels_to_EMU,
    TwoCellAnchor, AnchorMarker as SpreadshAnchor,
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.data_source import NumFmt as ChartNumFmt, StrRef as ChartStrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import column_index_from_string as col2idx
from dateutil.relativedelta import relativedelta

# ─────────────────────────────────────────────────────────────────────────────
# CORES
# ─────────────────────────────────────────────────────────────────────────────
C_PETROLEO = "175179"
C_CYAN     = "96CCD4"
C_GELO     = "E3F3FD"
C_ROXO     = "A64A8B"
C_CREME    = "F9F5CE"
C_AMBAR    = "B8860B"
C_WHITE    = "FFFFFF"
C_DARK     = "000000"

# ─────────────────────────────────────────────────────────────────────────────
# LAYOUT — colunas fixas (§11.4)
# ─────────────────────────────────────────────────────────────────────────────
VAL_COLS = ['C','E','G','I','K','M','O','Q','S','U','W','Y','AA']
PCT_COLS = ['D','F','H','J','L','N','P','R','T','V','X','Z','AB']
ACUM_V, ACUM_P  = 'AD', 'AE'
ROLL_V, ROLL_P  = 'AG', 'AH'
ANO_A,  ANO_B   = 'AJ', 'AK'
VAR_PCT, VAR_ABS = 'AL', 'AM'

# ─── Layout DFC (sem % AV, sem colunas de agregação — apenas 13 meses) ───────
DFC_VAL = ['C','D','E','F','G','H','I','J','K','L','M','N','O']  # 13 consecutivas
# sep cols DFC: A

DFC_CC = {
    'val':     DFC_VAL, 'pct':     [],
    'acum':    None,    'acum_p':  None,
    'roll':    None,    'roll_p':  None,
    'ano_a':   None,    'ano_b':   None,
    'var_pct': None,    'var_abs': None,
    'acum_fn': None,    'roll_fn': None,
    'varp_fn': None,    'vara_fn': None,
}

SEL_BU_ROW   = 4
SEL_TIPO_ROW = 5
SEL_ANO_ROW  = 6
HDR_ROW      = 7
DATA_ROW     = 8   # primeira linha de dados (sem sep antes do 1º N1)

H_SEP = 7.5
H_STD = 14.25
H_KPI = 15.75

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────────────────────────
def _fill(hex6):       return PatternFill("solid", fgColor=hex6)
def _font(c, b, s):    return Font(color=c, bold=b, size=s)
def _align(ind=0):     return Alignment(indent=ind)

def _sty_n1(c):   c.fill = _fill(C_CYAN);     c.font = _font(C_DARK,  True,  10)
def _sty_n2(c):   c.fill = _fill(C_GELO);     c.font = _font(C_DARK,  False, 10)
def _sty_n3(c):                                c.font = _font(C_DARK,  False, 10)
def _sty_kpi(c):  c.fill = _fill(C_PETROLEO); c.font = _font(C_WHITE, True,  11)
def _sty_roxo(c): c.fill = _fill(C_ROXO);     c.font = _font(C_WHITE, True,  11)
def _sty_sel(c):  c.fill = _fill(C_CREME);    c.font = _font(C_PETROLEO, True, 10)
def _sty_hdr(c):  c.fill = _fill(C_PETROLEO); c.font = _font(C_WHITE, True,  10); c.alignment = Alignment(horizontal="center")

def _row_sty(ws, row, sty_fn, cols):
    for col in cols:
        sty_fn(ws.cell(row=row, column=col2idx(col) if isinstance(col, str) else col))


def _abs_ref(cell_ref: str) -> str:
    """'AJ5' → '$AJ$5'"""
    m = re.match(r'([A-Z]+)(\d+)', cell_ref)
    return f"${m.group(1)}${m.group(2)}"


# ─────────────────────────────────────────────────────────────────────────────
# FORMATOS NUMÉRICOS E BORDAS (DesignDoc_Relatorio_v3 §Formatação)
# ─────────────────────────────────────────────────────────────────────────────
FMT_MOEDA = "#,##0.00;-#,##0.00"
FMT_PCT   = "#,##0.0%_);-#,##0.0%"

_SIDE_CYAN = Side(style="thin", color="96CCD4")
_SIDE_NONE = Side(style=None)
_BORDER_ALL_CYAN = Border(
    left=_SIDE_CYAN, right=_SIDE_CYAN,
    top=_SIDE_CYAN,  bottom=_SIDE_CYAN,
)


def _border_single(ws, ref):
    ws[ref].border = _BORDER_ALL_CYAN


def _border_merge_h(ws, col_start, col_end, row):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = Border(
            left  =_SIDE_CYAN if col == col_start else _SIDE_NONE,
            right =_SIDE_CYAN if col == col_end   else _SIDE_NONE,
            top   =_SIDE_CYAN,
            bottom=_SIDE_CYAN,
        )


def _apply_num_formats(ws, row, cc=None):
    _vc = (cc or {}).get('val',     VAL_COLS)
    _av = (cc or {}).get('acum',    ACUM_V)
    _rv = (cc or {}).get('roll',    ROLL_V)
    _aa = (cc or {}).get('ano_a',   ANO_A)
    _ab = (cc or {}).get('ano_b',   ANO_B)
    _va = (cc or {}).get('var_abs', VAR_ABS)
    _pc = (cc or {}).get('pct',     PCT_COLS)
    _ap = (cc or {}).get('acum_p',  ACUM_P)
    _rp = (cc or {}).get('roll_p',  ROLL_P)
    _vp = (cc or {}).get('var_pct', VAR_PCT)
    for col in [c for c in _vc + [_av, _rv, _aa, _ab, _va] if c]:
        ws.cell(row=row, column=col2idx(col)).number_format = FMT_MOEDA
    for col in _pc + [c for c in [_ap, _rp, _vp] if c]:
        ws.cell(row=row, column=col2idx(col)).number_format = FMT_PCT


_SEP_COLS  = {'A', 'AC', 'AF', 'AI', 'AN'}
_WIDTH_SEP = 1.86

_COL_A_PX  = 18   # col A = 1.86 unidades Excel = 18 px
_ROW_1_PX  = 20   # row 1 padrão = 15 pt = 20 px
_LOGO_H_PX = 25
_LOGO_H_PT = round(_LOGO_H_PX * 0.75, 2)  # 25 px → 18.75 pt


def _add_logo(ws, logo_path):
    """Logotipo com âncora absoluta em B2 — não move, não redimensiona."""
    if not logo_path:
        return
    try:
        img = XlImage(str(logo_path))
    except Exception:
        return
    orig_w, orig_h = img.width, img.height
    new_w = int(orig_w * _LOGO_H_PX / orig_h) if orig_h else _LOGO_H_PX
    img.width  = new_w
    img.height = _LOGO_H_PX
    img.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(pixels_to_EMU(_COL_A_PX), pixels_to_EMU(_ROW_1_PX)),
        ext=XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(_LOGO_H_PX)),
    )
    ws.add_image(img)
    ws.row_dimensions[2].height = _LOGO_H_PT


def _set_column_widths(ws):
    for col in _SEP_COLS:
        ws.column_dimensions[col].width = _WIDTH_SEP
    for col in VAL_COLS + [ACUM_V, ROLL_V, ANO_A, ANO_B, VAR_ABS]:
        ws.column_dimensions[col].width = 11
    for col in PCT_COLS + [ACUM_P, ROLL_P, VAR_PCT]:
        ws.column_dimensions[col].width = 7
    max_b = max(
        (len(str(ws.cell(row=r, column=2).value or ''))
         for r in range(1, ws.max_row + 1)),
        default=20
    )
    ws.column_dimensions['B'].width = min(max_b + 2, 50)


# ─────────────────────────────────────────────────────────────────────────────
# NAMED RANGES
# ─────────────────────────────────────────────────────────────────────────────
def _dn(wb, name, attr_text):
    if name not in wb.defined_names:
        wb.defined_names.add(DefinedName(name, attr_text=attr_text))


def _sheet_ref(sheet: str, cell: str) -> str:
    sn = f"'{sheet}'" if ' ' in sheet else sheet
    return f"{sn}!{_abs_ref(cell)}"


def _sheet_range_ref(sheet: str, cell1: str, cell2: str) -> str:
    sn = f"'{sheet}'" if ' ' in sheet else sheet
    return f"{sn}!{_abs_ref(cell1)}:{_abs_ref(cell2)}"


# ─────────────────────────────────────────────────────────────────────────────
# FÓRMULAS
# ─────────────────────────────────────────────────────────────────────────────
def _n3_val(col, row, n3_field="dre_n3"):
    return (
        f"=SUMIFS(f_Base[valor],"
        f"f_Base[mes_caixa],{col}${HDR_ROW},"
        f"f_Base[{n3_field}],$B{row},"
        f"f_Base[bu],IF(sel_BU=\"Todas\",\"*\",sel_BU),"
        f"f_Base[tipo_registro],"
        f"IF({col}${HDR_ROW}<=sel_MesCorte,\"Realizado\",sel_Projecao))"
    )

def _dfc_n1_total(col, n1_name):
    return (
        f"=SUMIFS(f_Base[valor],"
        f"f_Base[mes_caixa],{col}${HDR_ROW},"
        f"f_Base[dfc_n1],\"{n1_name}\","
        f"f_Base[bu],IF(sel_BU=\"Todas\",\"*\",sel_BU),"
        f"f_Base[tipo_registro],"
        f"IF({col}${HDR_ROW}<=sel_MesCorte,\"Realizado\",sel_Projecao))"
    )


def _sum_ch(col, rows):
    return f"=SUM({','.join(f'{col}{r}' for r in rows)})"

def _kpi_f(col, prev_row, n1_rows):
    refs = ",".join(f"{col}{r}" for r in n1_rows)
    return f"={col}{prev_row}+SUM({refs})" if prev_row else f"=SUM({refs})"

def _pct(val_col, row, ref_row):
    return f"={val_col}{row}/{val_col}${ref_row}"

def _acum(row):
    return f"=SUM({','.join(f'{c}{row}' for c in VAL_COLS)})"

def _rolling(row):
    c0, c1 = VAL_COLS[0], VAL_COLS[-1]
    return (
        f"=SUMPRODUCT("
        f"(COLUMN({c0}{row}:{c1}{row})>=29-2*sel_RollingN)*"
        f"(MOD(COLUMN({c0}{row}:{c1}{row})-3,2)=0)*"
        f"{c0}{row}:{c1}{row})"
    )

def _ano(ano_sel, tipo_sel, per_sel, row, n3_field="dre_n3"):
    return (
        f"=SUMIFS(f_Base[valor],"
        f"f_Base[ano],{ano_sel},"
        f"f_Base[{n3_field}],$B{row},"
        f"f_Base[bu],IF(sel_BU=\"Todas\",\"*\",sel_BU),"
        f"f_Base[tipo_registro],{tipo_sel},"
        f"IF(ISNUMBER(SEARCH(\"Trim\",{per_sel})),f_Base[trimestre],"
        f"IF(ISNUMBER(SEARCH(\"Sem\",{per_sel})),f_Base[semestre],"
        f"f_Base[mes_num])),"
        f"IF(ISNUMBER(SEARCH(\"Trim\",{per_sel})),VALUE(LEFT({per_sel},1)),"
        f"IF(ISNUMBER(SEARCH(\"Sem\",{per_sel})),VALUE(LEFT({per_sel},1)),"
        f"MATCH({per_sel},lista_periodo,0))))"
    )

def _var_pct_f(row):
    return f"=IF({ANO_B}{row}=0,\"\",({ANO_A}{row}-{ANO_B}{row})/ABS({ANO_B}{row}))"

def _var_abs_f(row):
    return f"={ANO_A}{row}-{ANO_B}{row}"

def _caixa_ini_first(col):
    return (
        f"=SUMIFS(f_SaldoBancos[valor],"
        f"f_SaldoBancos[data],EOMONTH({col}${HDR_ROW},-1),"
        f"f_SaldoBancos[BU],IF(sel_BU=\"Todas\",\"*\",sel_BU))"
    )


# ─────────────────────────────────────────────────────────────────────────────
# SELETORES — escrita compartilhada DRE/DFC
# ─────────────────────────────────────────────────────────────────────────────
def _write_selectors(ws, wb, mes_corte=None):
    sn = ws.title
    ws.sheet_view.showGridLines = False

    def dv_add(formula1, *cells):
        d = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(d)
        for c in cells:
            d.add(c)

    # ── Rótulos estáticos (B4, B5, E5) ───────────────────────────────────────
    for ref, label in [("B4", "Unidade"), ("B5", "Mês/Ano"), ("E5", "Projeção")]:
        ws[ref].value = label
        _border_single(ws, ref)

    # ── Linha 4: sel_BU (C4:G4 mesclado) ────────────────────────────────────
    ws.merge_cells("C4:G4")
    c = ws["C4"]
    c.value = "Todas"
    _sty_sel(c)
    c.alignment = Alignment(horizontal="left")
    _border_merge_h(ws, col2idx("C"), col2idx("G"), 4)
    dv_add("lista_bu", "C4")
    _dn(wb, "sel_BU", _sheet_ref(sn, "C4"))

    # ── Linha 5: sel_Ancora (C5) ─────────────────────────────────────────────
    _sty_sel(ws["C5"])
    if mes_corte:
        ws["C5"].value = mes_corte
    ws["C5"].number_format = "MMM/YYYY"
    ws["C5"].alignment = Alignment(horizontal="center")
    _border_single(ws, "C5")
    dv_add("lista_ancora", "C5")
    _dn(wb, "sel_Ancora", _sheet_ref(sn, "C5"))

    # ── Linha 5: sel_Projecao (F5:G5 mesclado) ───────────────────────────────
    ws.merge_cells("F5:G5")
    c = ws["F5"]
    c.value = "Orçado"
    _sty_sel(c)
    _border_merge_h(ws, col2idx("F"), col2idx("G"), 5)
    dv_add("lista_projecao", "F5")
    _dn(wb, "sel_Projecao", _sheet_ref(sn, "F5"))

    # ── Linha 5: sel_TipoA, sel_TipoB ────────────────────────────────────────
    for ref, dn_name in [("AJ5", "sel_TipoA"), ("AK5", "sel_TipoB")]:
        _sty_sel(ws[ref])
        _border_single(ws, ref)
        ws[ref].value     = "Realizado"
        ws[ref].alignment = Alignment(horizontal="center")
        dv_add("lista_tipo_registro", ref)
        _dn(wb, dn_name, _sheet_ref(sn, ref))

    # ── Linha 6: sel_AnoA, sel_AnoB ─────────────────────────────────────────
    for ref, dn_name, yr_offset in [
        ("AJ6", "sel_AnoA", -1),
        ("AK6", "sel_AnoB",  0),
    ]:
        _sty_sel(ws[ref])
        _border_single(ws, ref)
        if mes_corte:
            ws[ref].value = mes_corte.year + yr_offset
        ws[ref].alignment = Alignment(horizontal="center")
        dv_add("lista_anos", ref)
        _dn(wb, dn_name, _sheet_ref(sn, ref))

    # ── Linha 7: datas (VAL_COLS) — formato MMM/YYYY ─────────────────────────
    ws["AA7"] = "=sel_Ancora"
    _sty_hdr(ws["AA7"])
    ws["AA7"].number_format = "MMM/YYYY"
    prev = "AA"
    for col in reversed(VAL_COLS[:-1]):
        ws[f"{col}7"] = f"=EDATE({prev}7,-1)"
        _sty_hdr(ws[f"{col}7"])
        ws[f"{col}7"].number_format = "MMM/YYYY"
        prev = col

    for col in PCT_COLS:
        ws[f"{col}7"] = "% AV"
        _sty_hdr(ws[f"{col}7"])

    for cell_ref, label in [
        (f"{ACUM_V}7", "Acumulado"), (f"{ACUM_P}7", "% AV"),
        (f"{ROLL_P}7", "% AV"),
        (f"{VAR_PCT}7", "▲%"),      (f"{VAR_ABS}7", "▲R$"),
    ]:
        ws[cell_ref] = label
        _sty_hdr(ws[cell_ref])

    # ── sel_RollingN (AG7) — seletor (creme + borda), não cabeçalho ──────────
    ws["AG7"] = 6
    _sty_sel(ws["AG7"])
    _border_single(ws, "AG7")
    ws["AG7"].alignment   = Alignment(horizontal="center")
    ws["AG7"].number_format = "00"
    dv_add("lista_rolling_n", "AG7")
    _dn(wb, "sel_RollingN", _sheet_ref(sn, "AG7"))

    # ── sel_PeriodoA/B (AJ7/AK7) ─────────────────────────────────────────────
    periodo_default = _MESES[mes_corte.month - 1] if mes_corte else None
    for ref, dn_name in [("AJ7", "sel_PeriodoA"), ("AK7", "sel_PeriodoB")]:
        _sty_sel(ws[ref])
        _border_single(ws, ref)
        if periodo_default:
            ws[ref].value = periodo_default
        ws[ref].alignment = Alignment(horizontal="center")
        dv_add("lista_periodo", ref)
        _dn(wb, dn_name, _sheet_ref(sn, ref))

    for col in [VAR_PCT, VAR_ABS]:
        _sty_hdr(ws[f"{col}7"])

    # Formatação condicional: cabeçalho âmbar para meses projetados
    rule = FormulaRule(
        formula=[f"C${HDR_ROW}>sel_MesCorte"],
        fill=_fill(C_AMBAR),
        font=_font(C_WHITE, True, 10),
    )
    ws.conditional_formatting.add(f"C{HDR_ROW}:AA{HDR_ROW}", rule)


# ─────────────────────────────────────────────────────────────────────────────
# CONTRATO RECEITA BRUTA — regra universal AZ (§11.8.1)
# ─────────────────────────────────────────────────────────────────────────────
_N1_SUPRIMIDO = "Receita Líquida"
_N2_PROMOVIDO = "Receita Bruta"


def _validate_dre_contract(mapa_df):
    rl = mapa_df[mapa_df["dre_n1"] == _N1_SUPRIMIDO].sort_values("dre_ordem")
    if rl.empty:
        raise ValueError(
            f"MapaAloc inválido: dre_n1 '{_N1_SUPRIMIDO}' ausente. "
            f"O primeiro N1 do DRE deve ser '{_N1_SUPRIMIDO}'."
        )
    first_n2 = rl.iloc[0]["dre_n2"]
    if first_n2 != _N2_PROMOVIDO:
        raise ValueError(
            f"MapaAloc inválido: primeiro dre_n2 de '{_N1_SUPRIMIDO}' deve ser "
            f"'{_N2_PROMOVIDO}', encontrado {first_n2!r}."
        )


# ─────────────────────────────────────────────────────────────────────────────
# HIERARQUIA DO MAPAALOC
# ─────────────────────────────────────────────────────────────────────────────
def _hierarchy(mapa_df, n1_col, n2_col, n3_col, ord_col, excl=None):
    """
    Retorna {n1: {n2: [(min_ordem, n3)]}} com N3 deduplicados.
    Múltiplas categorias com o mesmo N3 colapsam em uma única linha de DRE/DFC
    (o SUMIFS referencia o label N3, somando todas as categorias daquele N3).
    """
    df = mapa_df.copy()
    if excl:
        df = df[~df[n1_col].isin(excl)]

    # Agrega: para cada (n1, n2, n3) mantém o menor dre/dfc_ordem
    grouped = (df.groupby([n1_col, n2_col, n3_col])[ord_col]
               .min()
               .reset_index())

    h = {}
    for _, r in grouped.iterrows():
        n1, n2, n3, o = r[n1_col], r[n2_col], r[n3_col], r[ord_col]
        h.setdefault(n1, {}).setdefault(n2, []).append((o, n3))

    for n1 in h:
        for n2 in h[n1]:
            h[n1][n2].sort()
        h[n1] = dict(sorted(h[n1].items(), key=lambda x: h[n1][x[0]][0][0]))
    return h


def _row_items(h, cascade, start_row=DATA_ROW):
    """
    cascade: [(n1_names, kpi_label_or_None, is_roxo), ...]
    Se kpi_label é None, nenhum KPI é inserido após o grupo.
    Retorna lista de dicts com 'row' atribuído.
    """
    items = []
    is_first = True

    for group_n1s, kpi_label, is_roxo in cascade:
        n1_group = []

        for n1_name in group_n1s:
            if n1_name not in h:
                continue

            if n1_name == _N1_SUPRIMIDO:
                # §11.8.1 — N1 "Receita Líquida" suprimido como cabeçalho;
                # N2 "Receita Bruta" promovido à primeira linha com estilo N1.
                is_first = False
                for n2_name, n3_list in h[n1_name].items():
                    if n2_name == _N2_PROMOVIDO:
                        rb_item = {"type": "n2_as_n1", "label": n2_name,
                                   "n3_items": []}
                        items.append(rb_item)
                        for _, n3_name in n3_list:
                            n3_item = {"type": "n3", "label": n3_name}
                            items.append(n3_item)
                            rb_item["n3_items"].append(n3_item)
                        items.append({"type": "sep"})
                        n1_group.append(rb_item)
                    else:
                        n2_item = {"type": "n2", "label": n2_name,
                                   "n3_items": []}
                        items.append(n2_item)
                        for _, n3_name in n3_list:
                            n3_item = {"type": "n3", "label": n3_name}
                            items.append(n3_item)
                            n2_item["n3_items"].append(n3_item)
                        n1_group.append(n2_item)
            else:
                if not is_first:
                    items.append({"type": "sep"})
                is_first = False

                n1_item = {"type": "n1", "label": n1_name, "n2_items": []}
                items.append(n1_item)

                for n2_name, n3_list in h[n1_name].items():
                    n2_item = {"type": "n2", "label": n2_name, "n3_items": []}
                    items.append(n2_item)
                    n1_item["n2_items"].append(n2_item)
                    for _, n3_name in n3_list:
                        n3_item = {"type": "n3", "label": n3_name}
                        items.append(n3_item)
                        n2_item["n3_items"].append(n3_item)

                n1_group.append(n1_item)

        if kpi_label is not None:
            items.append({"type": "sep"})
            items.append({"type": "kpi", "label": kpi_label,
                          "is_roxo": is_roxo, "n1_items": n1_group})

    # Atribuir números de linha
    row = start_row
    for item in items:
        item["row"] = row
        row += 1

    return items


def _link_kpi(items):
    prev = None
    for it in items:
        if it["type"] == "kpi":
            it["prev_kpi_row"] = prev
            prev = it["row"]


# ─────────────────────────────────────────────────────────────────────────────
# ESCRITA DE LINHAS
# ─────────────────────────────────────────────────────────────────────────────
def _write_items(ws, items, ref_row, n3_field="dre_n3", write_pct=True, cc=None):
    """Escreve todos os items na aba. ref_row = denominador %AV; cc=DFC_CC usa layout DFC."""
    _vc      = (cc or {}).get('val',     VAL_COLS)
    _pc      = (cc or {}).get('pct',     PCT_COLS)
    _av      = (cc or {}).get('acum',    ACUM_V)
    _ap      = (cc or {}).get('acum_p',  ACUM_P)
    _rv      = (cc or {}).get('roll',    ROLL_V)
    _rp      = (cc or {}).get('roll_p',  ROLL_P)
    _aa      = (cc or {}).get('ano_a',   ANO_A)
    _ab      = (cc or {}).get('ano_b',   ANO_B)
    _acum_fn = (cc or {}).get('acum_fn', _acum)
    _roll_fn = (cc or {}).get('roll_fn', _rolling)
    _sty_cols = _vc + _pc + [c for c in [_av, _ap, _rv, _rp, _aa, _ab] if c]
    for it in items:
        row = it["row"]
        t   = it["type"]

        if t == "sep":
            ws.row_dimensions[row].height = H_SEP
            continue

        b = ws.cell(row=row, column=2)

        if t == "n3":
            b.value = it["label"]
            b.alignment = _align(4)
            _sty_n3(b)
            ws.row_dimensions[row].height = H_STD
            ws.row_dimensions[row].outline_level = 1
            for col in _vc:
                ws.cell(row=row, column=col2idx(col)).value = _n3_val(col, row, n3_field)
            if _av and _acum_fn:
                ws.cell(row=row, column=col2idx(_av)).value = _acum_fn(row)
            if _rv and _roll_fn:
                ws.cell(row=row, column=col2idx(_rv)).value = _roll_fn(row)
            if _aa:
                ws.cell(row=row, column=col2idx(_aa)).value = _ano("sel_AnoA","sel_TipoA","sel_PeriodoA",row,n3_field)
            if _ab:
                ws.cell(row=row, column=col2idx(_ab)).value = _ano("sel_AnoB","sel_TipoB","sel_PeriodoB",row,n3_field)
            _pct_var(ws, row, ref_row, write_pct, cc)
            _row_sty(ws, row, _sty_n3, _sty_cols)

        elif t == "n2":
            b.value = it["label"]
            b.alignment = _align(2)
            _sty_n2(b)
            ws.row_dimensions[row].height = H_STD
            ch = [i["row"] for i in it["n3_items"]]
            for col in [c for c in _vc + [_av, _rv, _aa, _ab] if c]:
                ws.cell(row=row, column=col2idx(col)).value = _sum_ch(col, ch)
            _pct_var(ws, row, ref_row, write_pct, cc)
            _row_sty(ws, row, _sty_n2, _sty_cols)

        elif t == "n2_as_n1":
            b.value = it["label"]
            b.alignment = _align(0)
            _sty_n1(b)
            ws.row_dimensions[row].height = H_STD
            ch = [i["row"] for i in it["n3_items"]]
            for col in [c for c in _vc + [_av, _rv, _aa, _ab] if c]:
                ws.cell(row=row, column=col2idx(col)).value = _sum_ch(col, ch)
            _pct_var(ws, row, ref_row, write_pct, cc)
            _row_sty(ws, row, _sty_n1, _sty_cols)

        elif t == "n1":
            b.value = it["label"]
            b.alignment = _align(0)
            _sty_n1(b)
            ws.row_dimensions[row].height = H_STD
            ch = [i["row"] for i in it["n2_items"]]
            for col in [c for c in _vc + [_av, _rv, _aa, _ab] if c]:
                ws.cell(row=row, column=col2idx(col)).value = _sum_ch(col, ch)
            _pct_var(ws, row, ref_row, write_pct, cc)
            _row_sty(ws, row, _sty_n1, _sty_cols)

        elif t == "kpi":
            b.value = it["label"]
            b.alignment = _align(0)
            sty = _sty_roxo if it["is_roxo"] else _sty_kpi
            sty(b)
            ws.row_dimensions[row].height = H_KPI
            prev = it.get("prev_kpi_row")
            ch   = [i["row"] for i in it["n1_items"]]
            for col in [c for c in _vc + [_av, _rv, _aa, _ab] if c]:
                ws.cell(row=row, column=col2idx(col)).value = _kpi_f(col, prev, ch)
            _pct_var(ws, row, ref_row, write_pct, cc)
            _row_sty(ws, row, sty, _sty_cols)


def _pct_var(ws, row, ref_row, write_pct=True, cc=None):
    _pc    = (cc or {}).get('pct',     PCT_COLS)
    _av    = (cc or {}).get('val',     VAL_COLS)
    _avv   = (cc or {}).get('acum',    ACUM_V)
    _ap    = (cc or {}).get('acum_p',  ACUM_P)
    _rv    = (cc or {}).get('roll',    ROLL_V)
    _rp    = (cc or {}).get('roll_p',  ROLL_P)
    _vp    = (cc or {}).get('var_pct', VAR_PCT)
    _va    = (cc or {}).get('var_abs', VAR_ABS)
    _vpfn  = (cc or {}).get('varp_fn', _var_pct_f)
    _vafn  = (cc or {}).get('vara_fn', _var_abs_f)
    if write_pct and _pc:
        for i, p in enumerate(_pc):
            ws.cell(row=row, column=col2idx(p)).value = _pct(_av[i], row, ref_row)
        if _ap: ws.cell(row=row, column=col2idx(_ap)).value = _pct(_avv, row, ref_row)
        if _rp: ws.cell(row=row, column=col2idx(_rp)).value = _pct(_rv, row, ref_row)
    # VAR sempre escrita (variação Ano A/B, não % AV)
    if _vp and _vpfn:
        ws.cell(row=row, column=col2idx(_vp)).value = _vpfn(row)
    if _va and _vafn:
        ws.cell(row=row, column=col2idx(_va)).value = _vafn(row)
    _apply_num_formats(ws, row, cc)


def _apply_grouping(ws, items, group_pct_cols=True):
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    for it in items:
        if it["type"] == "n3":
            ws.row_dimensions[it["row"]].outline_level = 1
    if group_pct_cols:
        for col in PCT_COLS + [ACUM_P, ROLL_P]:
            ws.column_dimensions[col].outline_level = 1


# ─────────────────────────────────────────────────────────────────────────────
# KPIs — SELETORES INDEPENDENTES + TABELA + GRÁFICO
# ─────────────────────────────────────────────────────────────────────────────
_KPI_VAL_COLS = ['C','D','E','F','G','H','I','J','K','L','M','N','O']  # 13 meses


def _write_selectors_kpi(ws, wb, mes_corte):
    """Seletores independentes da aba KPIs — named ranges com prefixo kpi_."""
    sn = ws.title
    ws.sheet_view.showGridLines = False

    def dv_add(formula1, *cells):
        d = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(d)
        for c in cells:
            d.add(c)

    # Rótulos estáticos
    for ref, label in [("B4", "Unidade"), ("B5", "Mês/Ano"), ("E5", "Projeção")]:
        ws[ref].value = label
        _border_single(ws, ref)

    # kpi_sel_BU (C4:G4 mesclado)
    ws.merge_cells("C4:G4")
    c = ws["C4"]
    c.value = "Todas"
    _sty_sel(c)
    c.alignment = Alignment(horizontal="left")
    _border_merge_h(ws, col2idx("C"), col2idx("G"), 4)
    dv_add("lista_bu", "C4")
    _dn(wb, "kpi_sel_BU", _sheet_ref(sn, "C4"))

    # kpi_sel_Ancora (C5)
    _sty_sel(ws["C5"])
    if mes_corte:
        ws["C5"].value = mes_corte
    ws["C5"].number_format = "MMM/YYYY"
    ws["C5"].alignment = Alignment(horizontal="center")
    _border_single(ws, "C5")
    dv_add("lista_ancora", "C5")
    _dn(wb, "kpi_sel_Ancora", _sheet_ref(sn, "C5"))

    # kpi_sel_Projecao (F5:G5 mesclado)
    ws.merge_cells("F5:G5")
    c = ws["F5"]
    c.value = "Orçado"
    _sty_sel(c)
    _border_merge_h(ws, col2idx("F"), col2idx("G"), 5)
    dv_add("lista_projecao", "F5")
    _dn(wb, "kpi_sel_Projecao", _sheet_ref(sn, "F5"))

    # Row 7: cabeçalhos de data (C-O consecutivas, como DFC)
    ws[f"O{HDR_ROW}"] = "=kpi_sel_Ancora"
    _sty_hdr(ws[f"O{HDR_ROW}"])
    ws[f"O{HDR_ROW}"].number_format = "MMM/YYYY"
    prev = "O"
    for col in reversed(_KPI_VAL_COLS[:-1]):
        ws[f"{col}{HDR_ROW}"] = f"=EDATE({prev}{HDR_ROW},-1)"
        _sty_hdr(ws[f"{col}{HDR_ROW}"])
        ws[f"{col}{HDR_ROW}"].number_format = "MMM/YYYY"
        prev = col

    # Formatação condicional: âmbar para meses projetados
    rule = FormulaRule(
        formula=[f"C${HDR_ROW}>sel_MesCorte"],
        fill=_fill(C_AMBAR),
        font=_font(C_WHITE, True, 10),
    )
    ws.conditional_formatting.add(f"C{HDR_ROW}:O{HDR_ROW}", rule)


def build_kpi(wb, mes_corte=None, logo_path=None):
    ws = wb["KPIs"]
    _add_logo(ws, logo_path)
    _write_selectors_kpi(ws, wb, mes_corte)

    # ── Tabela: Receita Líquida (linha DATA_ROW) ──────────────────────────────
    b = ws.cell(row=DATA_ROW, column=2, value="Receita Líquida")
    b.font = Font(bold=True, size=10, color=C_DARK)
    ws.row_dimensions[DATA_ROW].height = H_STD

    for col in _KPI_VAL_COLS:
        formula = (
            f"=SUMIFS(f_Base[valor],"
            f"f_Base[mes_caixa],{col}${HDR_ROW},"
            f"f_Base[kpi_receita_liquida],\"Sim\","
            f"f_Base[bu],IF(kpi_sel_BU=\"Todas\",\"*\",kpi_sel_BU),"
            f"f_Base[tipo_registro],"
            f"IF({col}${HDR_ROW}<=sel_MesCorte,\"Realizado\",kpi_sel_Projecao))"
        )
        cell = ws.cell(row=DATA_ROW, column=col2idx(col), value=formula)
        cell.number_format = FMT_MOEDA
        cell.font = Font(size=10, color=C_DARK)

    # ── Linha separadora ──────────────────────────────────────────────────────
    ws.row_dimensions[DATA_ROW + 1].height = H_SEP

    # ── Gráfico de colunas: Receita Líquida ───────────────────────────────────
    # Linhas 11-16: altura 60pt (~12.7cm total, proporção ~2:1 com largura de 26cm)
    for r in range(11, 17):
        ws.row_dimensions[r].height = 60

    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "clustered"
    # Sem título visível (autoTitleDeleted implícito)

    # Dados e categorias
    data_ref = Reference(ws, min_col=col2idx("C"), max_col=col2idx("O"),
                         min_row=DATA_ROW, max_row=DATA_ROW)
    chart.add_data(data_ref)
    cats_ref = Reference(ws, min_col=col2idx("C"), max_col=col2idx("O"),
                         min_row=HDR_ROW)
    chart.set_categories(cats_ref)

    s = chart.series[0]

    # Nome da série = célula B8 (dinâmico)
    s.tx = SeriesLabel(strRef=ChartStrRef(f='KPIs!$B$8'))

    # Cor: azul-petróleo fixo (opção B) — sem borda nas barras
    s.graphicalProperties.solidFill = C_PETROLEO
    s.invertIfNegative = False

    # Data labels: valores em milhares ("K"), 9pt negrito
    dlbls = DataLabelList()
    dlbls.showVal          = True
    dlbls.showCatName      = False
    dlbls.showSerName      = False
    dlbls.showPercent      = False
    dlbls.showLegendKey    = False
    dlbls.showLeaderLines  = False
    dlbls.numFmt = ChartNumFmt(formatCode='#,##0.0,"K"', sourceLinked=False)
    s.dLbls = dlbls

    # Legenda: topo, sem borda/fill
    chart.legend = Legend()
    chart.legend.legendPos = 't'
    chart.legend.overlay   = False

    # Posição: C11:O16 via TwoCellAnchor
    anchor = TwoCellAnchor()
    anchor._from = SpreadshAnchor(col=col2idx("C") - 1, row=10, colOff=0, rowOff=0)
    anchor.to    = SpreadshAnchor(col=col2idx("O"),     row=15, colOff=0, rowOff=0)
    ws.add_chart(chart)
    ws._charts[-1].anchor = anchor

    # ── Larguras de coluna ────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = _WIDTH_SEP
    for col in _KPI_VAL_COLS:
        ws.column_dimensions[col].width = 11
    ws.column_dimensions['B'].width = 22


# ─────────────────────────────────────────────────────────────────────────────
# DFC — HEADER E LARGURAS ESPECÍFICAS
# ─────────────────────────────────────────────────────────────────────────────
def _write_header_dfc(ws, mes_corte):
    """Header row 7 do DFC: 13 colunas mensais consecutivas C-O."""
    # Datas: O7 = mais recente (sel_Ancora), cadeia para trás até C7
    ws[f"O{HDR_ROW}"] = "=sel_Ancora"
    _sty_hdr(ws[f"O{HDR_ROW}"])
    ws[f"O{HDR_ROW}"].number_format = "MMM/YYYY"
    prev = "O"
    for col in reversed(DFC_VAL[:-1]):
        ws[f"{col}{HDR_ROW}"] = f"=EDATE({prev}{HDR_ROW},-1)"
        _sty_hdr(ws[f"{col}{HDR_ROW}"])
        ws[f"{col}{HDR_ROW}"].number_format = "MMM/YYYY"
        prev = col

    # Formatação condicional: âmbar para meses projetados
    rule = FormulaRule(
        formula=[f"C${HDR_ROW}>sel_MesCorte"],
        fill=_fill(C_AMBAR),
        font=_font(C_WHITE, True, 10),
    )
    ws.conditional_formatting.add(f"C{HDR_ROW}:O{HDR_ROW}", rule)


def _set_column_widths_dfc(ws):
    ws.column_dimensions['A'].width = _WIDTH_SEP
    for col in DFC_VAL:
        ws.column_dimensions[col].width = 11
    max_b = max(
        (len(str(ws.cell(row=r, column=2).value or ''))
         for r in range(1, ws.max_row + 1)),
        default=20
    )
    ws.column_dimensions['B'].width = min(max_b + 2, 50)


# ─────────────────────────────────────────────────────────────────────────────
# BUILD LISTA
# ─────────────────────────────────────────────────────────────────────────────
_MESES = [
    "janeiro","fevereiro","março","abril","maio","junho",
    "julho","agosto","setembro","outubro","novembro","dezembro",
]
_TRIMS = ["1º Trim","2º Trim","3º Trim","4º Trim"]
_SEMS  = ["1º Sem","2º Sem"]


def build_lista(wb, f_base_df, bu_validos):
    ws = wb["Lista"]

    def _write_col(col_idx, values, dn_name):
        for i, v in enumerate(values, 1):
            ws.cell(row=i, column=col_idx, value=v)
        last = f"${chr(64 + col_idx)}${len(values)}" if col_idx <= 26 else None
        if col_idx <= 26:
            cell1 = f"{chr(64+col_idx)}1"
            celln = f"{chr(64+col_idx)}{len(values)}"
        else:
            from openpyxl.utils import get_column_letter
            l = get_column_letter(col_idx)
            cell1, celln = f"{l}1", f"{l}{len(values)}"
        _dn(wb, dn_name, _sheet_range_ref("Lista", cell1, celln))

    _write_col(1, _MESES + _TRIMS + _SEMS, "lista_periodo")
    _write_col(2, list(range(2, 13)),        "lista_rolling_n")
    _write_col(3, ["Orçado", "Reforecast"],  "lista_projecao")
    _write_col(4, ["Realizado","Orçado","Reforecast"], "lista_tipo_registro")
    _write_col(5, _ancora_dates(f_base_df),  "lista_ancora")
    _write_col(6, _anos(f_base_df),          "lista_anos")
    _write_col(7, list(bu_validos) + ["Todas"], "lista_bu")


def _ancora_dates(df):
    mc = pd.to_datetime(df["mes_caixa"])
    if mc.empty or mc.isna().all():
        return []
    first = mc.min().date().replace(day=1)
    start = first + relativedelta(months=12)
    orcado = df[df["tipo_registro"] == "Orçado"]["mes_caixa"]
    last_year = (pd.to_datetime(orcado).dt.year.max() if not orcado.empty
                 else mc.dt.year.max())
    from datetime import date
    end = date(last_year, 12, 1)
    dates, cur = [], start
    while cur <= end:
        dates.append(cur)
        cur += relativedelta(months=1)
    return dates


def _anos(df):
    mc = pd.to_datetime(df["mes_caixa"])
    if mc.empty or mc.isna().all():
        return []
    first_y = mc.dt.year.min()
    orcado = df[df["tipo_registro"] == "Orçado"]["mes_caixa"]
    last_y = (pd.to_datetime(orcado).dt.year.max() if not orcado.empty
              else mc.dt.year.max())
    return list(range(first_y, last_y + 1))


# ─────────────────────────────────────────────────────────────────────────────
# BUILD CAD_CLIENTE
# ─────────────────────────────────────────────────────────────────────────────
def build_cad_cliente(wb, config: dict):
    ws = wb["cad_cliente"]
    bold = Font(bold=True, size=10)

    groups = [
        ("Identificação",   ["codigo","nome","segmento_cliente","status"]),
        ("Origem de dados", ["origem_dados_realizado","path_lctos"]),
        ("Staging",         ["staging_mapa_fonte","conversao_defensiva_valor"]),
        ("BU",              ["bu_origem","bu_valores_validos"]),
        ("Colunas cond.",   ["tem_conta_bancaria","tem_fornecedor_cliente"]),
        ("Projeção",        ["mes_corte_realizado","reforecast_vigente_ref"]),
        ("MapaAloc",        ["mapaaloc_arquivo","mapaaloc_versao"]),
        ("Moeda",           ["moeda"]),
    ]

    row = 1
    mes_corte_row = None

    for grp, fields in groups:
        ws.cell(row=row, column=1, value=grp).font = bold
        row += 1
        for field in fields:
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=config.get(field, ""))
            if field == "mes_corte_realizado":
                mes_corte_row = row
            row += 1
        row += 1

    if mes_corte_row:
        _dn(wb, "cad_mes_corte", f"cad_cliente!$B${mes_corte_row}")
        ws.cell(row=row, column=1, value="sel_MesCorte (aux)")
        ws.cell(row=row, column=2,
                value="=DATE(VALUE(LEFT(cad_mes_corte,4)),"
                      "VALUE(MID(cad_mes_corte,6,2)),1)")
        _dn(wb, "sel_MesCorte", f"cad_cliente!$B${row}")


# ─────────────────────────────────────────────────────────────────────────────
# BUILD DRE
# ─────────────────────────────────────────────────────────────────────────────
def build_dre(wb, mapa_df, cascade, mes_corte=None, logo_path=None):
    """
    cascade: [(n1_names_list, kpi_label, is_roxo), ...]
    is_roxo=True → linha roxo-logo (RESULTADO INVESTIDORES)
    """
    _validate_dre_contract(mapa_df)
    ws = wb["DRE Gerencial"]
    _write_selectors(ws, wb, mes_corte)
    _add_logo(ws, logo_path)

    h = _hierarchy(mapa_df, "dre_n1","dre_n2","dre_n3","dre_ordem",
                   excl=["Efeito Zero"])
    items = _row_items(h, cascade, start_row=DATA_ROW)
    _link_kpi(items)

    rb_row = DATA_ROW  # Receita Bruta = primeira linha (sem sep antes)
    _write_items(ws, items, ref_row=rb_row, n3_field="dre_n3")
    _apply_grouping(ws, items)
    _set_column_widths(ws)


# ─────────────────────────────────────────────────────────────────────────────
# BUILD DFC
# ─────────────────────────────────────────────────────────────────────────────
_DFC_N1S = [
    "Atividades Operacionais",
    "Atividades Não Operacionais",
    "Atividades de Investimento",
    "Atividades de Financiamento",
]
_DFC_EXCL = ["SEM_DFC", "Efeito Zero"]


def build_dfc(wb, mapa_df, mes_corte=None, logo_path=None):
    ws = wb["DFC"]
    ws.sheet_view.showGridLines = False
    _add_logo(ws, logo_path)
    _write_header_dfc(ws, mes_corte)

    # ── Seção Resumo ─────────────────────────────────────────────────────────
    summ_n1_rows = {}
    for i, n1 in enumerate(_DFC_N1S):
        r = DATA_ROW + i
        c = ws.cell(row=r, column=2, value=n1)
        c.font = _font(C_DARK, True, 10)
        ws.row_dimensions[r].height = H_STD
        summ_n1_rows[n1] = r
        for col in DFC_VAL:
            ws.cell(row=r, column=col2idx(col)).value = _dfc_n1_total(col, n1)
        _row_sty(ws, r, _sty_n3, DFC_VAL)
        _apply_num_formats(ws, r, DFC_CC)

    caixa_ini_row = DATA_ROW + 4
    caixa_fim_row = DATA_ROW + 5
    fluxo_row     = DATA_ROW + 6

    # CAIXA INÍCIO
    c = ws.cell(row=caixa_ini_row, column=2, value="CAIXA - INÍCIO DO MÊS")
    _sty_n1(c)
    ws.row_dimensions[caixa_ini_row].height = H_STD
    for j, col in enumerate(DFC_VAL):
        if j == 0:
            ws.cell(row=caixa_ini_row, column=col2idx(col)).value = _caixa_ini_first(col)
        else:
            ws.cell(row=caixa_ini_row, column=col2idx(col)).value = f"={DFC_VAL[j-1]}{caixa_fim_row}"
    _row_sty(ws, caixa_ini_row, _sty_n1, DFC_VAL)
    _apply_num_formats(ws, caixa_ini_row, DFC_CC)

    # CAIXA FIM (ACUM vazio — estoque sem significado como acumulado)
    c = ws.cell(row=caixa_fim_row, column=2, value="CAIXA - FIM DO MÊS")
    _sty_n1(c)
    ws.row_dimensions[caixa_fim_row].height = H_STD
    for col in DFC_VAL:
        ws.cell(row=caixa_fim_row, column=col2idx(col)).value = (
            f"={col}{caixa_ini_row}+{col}{fluxo_row}"
        )
    _row_sty(ws, caixa_fim_row, _sty_n1, DFC_VAL)
    _apply_num_formats(ws, caixa_fim_row, DFC_CC)

    # FLUXO DE CAIXA
    c = ws.cell(row=fluxo_row, column=2, value="FLUXO DE CAIXA")
    _sty_kpi(c)
    ws.row_dimensions[fluxo_row].height = H_KPI
    s_rows = list(summ_n1_rows.values())
    for col in DFC_VAL:
        ws.cell(row=fluxo_row, column=col2idx(col)).value = _sum_ch(col, s_rows)
    _row_sty(ws, fluxo_row, _sty_kpi, DFC_VAL)
    _apply_num_formats(ws, fluxo_row, DFC_CC)

    # ── Separador após resumo ─────────────────────────────────────────────────
    sep_row = DATA_ROW + 7
    ws.row_dimensions[sep_row].height = H_SEP

    # ── Seção Detalhe ─────────────────────────────────────────────────────────
    h = _hierarchy(mapa_df, "dfc_n1","dfc_n2","dfc_n3","dfc_ordem", excl=_DFC_EXCL)
    cascade_dfc = [([n1], None, False) for n1 in _DFC_N1S if n1 in h]
    detail_start = sep_row + 1
    items = _row_items(h, cascade_dfc, start_row=detail_start)

    _write_items(ws, items, ref_row=fluxo_row, n3_field="dfc_n3",
                 write_pct=False, cc=DFC_CC)

    # Última linha: Movimentação Mês (roxo-logo — espelha FLUXO)
    data_rows  = [it["row"] for it in items if it["type"] != "sep"]
    last_data  = max(data_rows) if data_rows else detail_start
    mov_sep    = last_data + 1
    mov_row    = last_data + 2
    ws.row_dimensions[mov_sep].height = H_SEP

    c = ws.cell(row=mov_row, column=2, value="Movimentação Mês")
    _sty_roxo(c)
    ws.row_dimensions[mov_row].height = H_KPI
    for col in DFC_VAL:
        ws.cell(row=mov_row, column=col2idx(col)).value = f"={col}{fluxo_row}"
    _row_sty(ws, mov_row, _sty_roxo, DFC_VAL)
    _apply_num_formats(ws, mov_row, DFC_CC)

    _apply_grouping(ws, items, group_pct_cols=False)
    _set_column_widths_dfc(ws)
