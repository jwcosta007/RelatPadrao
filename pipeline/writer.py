import os
import zipfile
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path


def create_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def write_df_table(wb: openpyxl.Workbook, sheet_name: str,
                   df: pd.DataFrame, table_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    n_cols = len(df.columns)
    n_rows = max(len(df) + 1, 2)  # OpenXML exige header + ≥1 linha de dados
    ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)


def save(wb: openpyxl.Workbook, path: Path) -> None:
    wb.save(path)


# Template do chart KPIs v3: eixos visíveis, tickMarks só nas quebras principais
_KPI_CHART_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<c:chartSpace'
    ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
    ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
    '<c:date1904 val="0"/><c:lang val="pt-BR"/><c:roundedCorners val="0"/>'
    '<c:chart>'

    # Título
    '<c:title>'
    '<c:tx><c:rich>'
    '<a:bodyPr/><a:lstStyle/>'
    '<a:p><a:pPr><a:defRPr b="1" sz="1400"/></a:pPr>'
    '<a:r><a:rPr lang="pt-BR" b="1" sz="1400"/>'
    '<a:t>Receita Líquida</a:t>'
    '</a:r></a:p>'
    '</c:rich></c:tx>'
    '<c:overlay val="0"/>'
    '</c:title>'
    '<c:autoTitleDeleted val="0"/>'

    # Plot area
    '<c:plotArea>'
    '<c:barChart>'
    '<c:barDir val="col"/><c:grouping val="clustered"/>'
    '<c:varyColors val="0"/>'
    '<c:gapWidth val="100"/>'
    '<c:ser>'
    '<c:idx val="0"/><c:order val="0"/>'
    '<c:tx><c:strRef><c:f>KPIs!$B$8</c:f></c:strRef></c:tx>'
    '<c:spPr>'
    '<a:solidFill><a:srgbClr val="175179"/></a:solidFill>'
    '<a:ln><a:noFill/></a:ln>'
    '</c:spPr>'
    '<c:invertIfNegative val="0"/>'
    '<c:dLbls>'
    '<c:numFmt formatCode="#,##0.0,&quot;K&quot;" sourceLinked="0"/>'
    '<c:spPr><a:noFill/><a:ln><a:noFill/></a:ln></c:spPr>'
    '<c:txPr><a:bodyPr/><a:lstStyle/>'
    '<a:p><a:pPr><a:defRPr sz="900" b="1"/></a:pPr></a:p>'
    '</c:txPr>'
    '<c:showLegendKey val="0"/><c:showVal val="1"/>'
    '<c:showCatName val="0"/><c:showSerName val="0"/>'
    '<c:showPercent val="0"/><c:showBubbleSize val="0"/>'
    '<c:showLeaderLines val="0"/>'
    '</c:dLbls>'
    '<c:cat><c:numRef><c:f>KPIs!$C$7:$O$7</c:f></c:numRef></c:cat>'
    '<c:val><c:numRef><c:f>KPIs!$C$8:$O$8</c:f></c:numRef></c:val>'
    '</c:ser>'
    '<c:axId val="1"/><c:axId val="2"/>'
    '</c:barChart>'

    # plotArea: fundo transparente (sem spPr nos eixos)
    '<c:spPr><a:noFill/><a:ln><a:noFill/></a:ln></c:spPr>'

    # Eixo X — rótulos negrito 8pt, marcadores visíveis
    '<c:catAx>'
    '<c:axId val="1"/>'
    '<c:scaling><c:orientation val="minMax"/></c:scaling>'
    '<c:delete val="0"/><c:axPos val="b"/>'
    '<c:numFmt formatCode="mmm/yyyy" sourceLinked="0"/>'
    '<c:majorTickMark val="out"/>'
    '<c:minorTickMark val="none"/>'
    '<c:tickLblPos val="nextTo"/>'
    '<c:txPr><a:bodyPr/><a:lstStyle/>'
    '<a:p><a:pPr><a:defRPr b="1" sz="800"/></a:pPr></a:p>'
    '</c:txPr>'
    '<c:crossAx val="2"/>'
    '<c:crosses val="autoZero"/><c:auto val="1"/>'
    '<c:lblAlgn val="ctr"/><c:lblOffset val="100"/>'
    '</c:catAx>'

    # Eixo Y — 4-5 quebras (100K), só marcadores principais
    '<c:valAx>'
    '<c:axId val="2"/>'
    '<c:scaling>'
    '<c:orientation val="minMax"/>'
    '<c:min val="0"/>'
    '</c:scaling>'
    '<c:delete val="0"/><c:axPos val="l"/>'
    '<c:numFmt formatCode="#,##0,&quot;K&quot;" sourceLinked="0"/>'
    '<c:majorTickMark val="out"/>'
    '<c:minorTickMark val="none"/>'
    '<c:tickLblPos val="nextTo"/>'
    '<c:crossAx val="1"/>'
    '<c:crosses val="autoZero"/>'
    '<c:crossBetween val="between"/>'
    '<c:majorUnit val="100000"/>'
    '</c:valAx>'

    '</c:plotArea>'
    '<c:legend><c:legendPos val="t"/><c:overlay val="0"/></c:legend>'
    '<c:plotVisOnly val="1"/><c:dispBlanksAs val="gap"/>'
    '</c:chart>'

    # Borda externa cyan-pastel 0.75pt, fundo branco
    '<c:spPr>'
    '<a:solidFill><a:schemeClr val="bg1"/></a:solidFill>'
    '<a:ln w="9525">'
    '<a:solidFill><a:srgbClr val="96CCD4"/></a:solidFill>'
    '</a:ln>'
    '</c:spPr>'
    '</c:chartSpace>'
).encode('utf-8')




def patch_kpi_chart(path: Path) -> None:
    """Substitui chart1.xml pelo template correto (1 série, formatação completa)."""
    tmp = Path(str(path) + '.tmp')
    try:
        with zipfile.ZipFile(path, 'r') as zin:
            with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == 'xl/charts/chart1.xml':
                        data = _KPI_CHART_XML
                    zout.writestr(item, data)
        os.replace(tmp, path)
    except Exception:
        if tmp.exists():
            tmp.unlink(missing_ok=True)
